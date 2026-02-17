"""スプレッドシート出力モジュール

Excel(.xlsx)形式で出力（文字化けなし）:
1. Amazon商品画像URL
2. 商品ページURL
3. アリババショップURL
4. 販売価格（ライバル価格-10円）
5. アリババ原単価（円）
6. 利益率（%）
7. 利幅（円）
"""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from ..models.result import ResearchResult
from ..models.product import ProductDetail, AlibabaProduct
from ..models.result import ProfitResult
from ..config import get_config

logger = logging.getLogger(__name__)


class SpreadsheetExporter:
    """Excel形式(.xlsx)でリサーチ結果を出力

    ライバル価格 - 10円での利益計算を含む。
    """

    # ヘッダー
    HEADERS = [
        "No.",
        "Amazon商品画像",
        "商品ページURL",
        "1688ショップURL",
        "販売価格（円）",
        "アリババ原単価（円）",
        "利益率（%）",
        "利幅（円）",
    ]

    def __init__(
        self,
        output_dir: Optional[str | Path] = None,
        encoding: Optional[str] = None,
        price_adjustment: int = -10,
    ):
        """
        Args:
            output_dir: 出力ディレクトリ
            encoding: 未使用（後方互換のため残す）
            price_adjustment: ライバル価格からの調整（デフォルト: -10円）
        """
        if output_dir is None:
            base_dir = Path(__file__).parent.parent.parent
            output_dir = base_dir / "output" / "results"

        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.price_adjustment = price_adjustment

    def _style_header(self, ws, headers: list[str]) -> None:
        """ヘッダー行にスタイルを適用"""
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    def _auto_column_width(self, ws) -> None:
        """列幅を自動調整"""
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    # 日本語は2文字分の幅
                    val = str(cell.value)
                    length = sum(2 if ord(c) > 127 else 1 for c in val)
                    max_length = max(max_length, length)
            ws.column_dimensions[col_letter].width = min(max_length + 4, 60)

    def export(
        self,
        results: list[ResearchResult],
        keyword: str,
        filename: Optional[str] = None,
    ) -> Path:
        """リサーチ結果をExcel形式で出力

        Args:
            results: リサーチ結果リスト
            keyword: 検索キーワード（ファイル名に使用）
            filename: 出力ファイル名（省略時は自動生成）

        Returns:
            出力ファイルのパス
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_keyword = "".join(
                c for c in keyword if c.isalnum() or c in ('_', '-', ' ')
            ).strip().replace(' ', '_')
            filename = f"spreadsheet_{safe_keyword}_{timestamp}.xlsx"

        filepath = self.output_dir / filename

        wb = Workbook()
        ws = wb.active
        ws.title = "リサーチ結果"

        # ヘッダー
        self._style_header(ws, self.HEADERS)

        # データ行
        for i, result in enumerate(results, 1):
            row_data = self._create_row(i, result)
            row_idx = i + 1  # ヘッダーが1行目
            for col_idx, header in enumerate(self.HEADERS, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data[header])

        # 列幅調整
        self._auto_column_width(ws)

        wb.save(filepath)
        logger.info(f"スプレッドシート出力完了: {filepath} ({len(results)}件)")
        return filepath

    def _create_row(self, index: int, result: ResearchResult) -> dict:
        """1行のデータを作成"""
        amazon = result.amazon_product
        alibaba = result.alibaba_product
        profit = result.profit_result

        # 販売価格 = ライバル価格 - 10円
        selling_price = amazon.price + self.price_adjustment

        # アリババ原単価（円換算）
        config = get_config()
        exchange_rate = config.profit.exchange_rate
        alibaba_price_jpy = int(alibaba.price_cny * exchange_rate)

        # 利幅と利益率を再計算（販売価格-10円ベース）
        profit_margin = selling_price - profit.total_cost
        profit_rate = (profit_margin / selling_price * 100) if selling_price > 0 else 0

        return {
            "No.": index,
            "Amazon商品画像": amazon.image_url,
            "商品ページURL": amazon.product_url or f"https://www.amazon.co.jp/dp/{amazon.asin}",
            "1688ショップURL": alibaba.shop_url or "",
            "販売価格（円）": selling_price,
            "アリババ原単価（円）": alibaba_price_jpy,
            "利益率（%）": round(profit_rate, 1),
            "利幅（円）": profit_margin,
        }

    def export_detailed(
        self,
        results: list[ResearchResult],
        keyword: str,
        filename: Optional[str] = None,
    ) -> Path:
        """詳細版スプレッドシート出力（追加情報を含む）

        Args:
            results: リサーチ結果リスト
            keyword: 検索キーワード
            filename: 出力ファイル名

        Returns:
            出力ファイルのパス
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_keyword = "".join(
                c for c in keyword if c.isalnum() or c in ('_', '-', ' ')
            ).strip().replace(' ', '_')
            filename = f"detailed_{safe_keyword}_{timestamp}.xlsx"

        filepath = self.output_dir / filename

        detailed_headers = [
            "No.",
            "ASIN",
            "商品タイトル",
            "Amazon商品画像",
            "商品ページURL",
            "ライバル価格（円）",
            "販売価格（円）",
            "1688商品URL",
            "1688ショップURL",
            "アリババ原単価（元）",
            "アリババ原単価（円）",
            "総コスト（円）",
            "利幅（円）",
            "利益率（%）",
            "レビュー数",
            "評価",
            "BSR",
            "バリエーション数",
        ]

        config = get_config()
        exchange_rate = config.profit.exchange_rate

        wb = Workbook()
        ws = wb.active
        ws.title = "詳細リサーチ結果"

        # ヘッダー
        self._style_header(ws, detailed_headers)

        # データ行
        for i, result in enumerate(results, 1):
            amazon = result.amazon_product
            alibaba = result.alibaba_product
            profit = result.profit_result

            selling_price = amazon.price + self.price_adjustment
            alibaba_price_jpy = int(alibaba.price_cny * exchange_rate)
            profit_margin = selling_price - profit.total_cost
            profit_rate = (profit_margin / selling_price * 100) if selling_price > 0 else 0

            row_data = {
                "No.": i,
                "ASIN": amazon.asin,
                "商品タイトル": amazon.title[:50] + "..." if len(amazon.title) > 50 else amazon.title,
                "Amazon商品画像": amazon.image_url,
                "商品ページURL": amazon.product_url or f"https://www.amazon.co.jp/dp/{amazon.asin}",
                "ライバル価格（円）": amazon.price,
                "販売価格（円）": selling_price,
                "1688商品URL": alibaba.product_url,
                "1688ショップURL": alibaba.shop_url or "",
                "アリババ原単価（元）": alibaba.price_cny,
                "アリババ原単価（円）": alibaba_price_jpy,
                "総コスト（円）": profit.total_cost,
                "利幅（円）": profit_margin,
                "利益率（%）": round(profit_rate, 1),
                "レビュー数": amazon.review_count,
                "評価": amazon.rating or "",
                "BSR": amazon.bsr,
                "バリエーション数": amazon.variation_count,
            }

            row_idx = i + 1
            for col_idx, header in enumerate(detailed_headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data[header])

        # 列幅調整
        self._auto_column_width(ws)

        wb.save(filepath)
        logger.info(f"詳細スプレッドシート出力完了: {filepath} ({len(results)}件)")
        return filepath
