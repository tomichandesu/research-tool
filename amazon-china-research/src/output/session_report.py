"""セッション統合レポート生成モジュール

自動リサーチのセッション全体の結果を1つのHTML + 1つのExcelに統合する。
キーワードごとにバラバラだったレポートを1ファイルにまとめる。
"""
from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


class SessionReportGenerator:
    """セッション統合レポートジェネレーター"""

    def __init__(self, output_dir: str = "output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate_html(
        self,
        session_data: list[dict],
        stats: dict[str, Any],
    ) -> str:
        """統合HTMLレポートを生成

        Args:
            session_data: KWごとのデータリスト
                [{"keyword": str, "products": list[dict], "score": float,
                  "total_searched": int, "pass_count": int}, ...]
            stats: セッション統計
                {"total_keywords": int, "total_candidates": int,
                 "elapsed_seconds": float, "elapsed_str": str}

        Returns:
            生成したHTMLファイルのパス
        """
        # スコア降順でソート
        sorted_data = sorted(session_data, key=lambda x: -x.get("score", 0))

        data = {
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "stats": stats,
            "keywords": sorted_data,
        }

        # Use ensure_ascii=True to escape all special/control characters safely.
        # Embed as a direct JS object literal (no JSON.parse + string escaping needed).
        json_data = json.dumps(data, ensure_ascii=True)
        html = _SESSION_HTML_TEMPLATE.replace("__DATA_PLACEHOLDER__", json_data)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = self.output_dir / f"session_{timestamp}.html"

        with open(filepath, "w", encoding="utf-8") as f:
            f.write(html)

        logger.info(f"セッション統合HTML生成: {filepath}")
        return str(filepath)

    def generate_excel(
        self,
        session_data: list[dict],
        stats: dict[str, Any],
    ) -> str:
        """統合Excelレポートを生成

        Sheet1: キーワードサマリー
        Sheet2: 全商品詳細

        Args:
            session_data: KWごとのデータリスト
            stats: セッション統計

        Returns:
            生成したExcelファイルのパス
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()

        # --- Sheet1: キーワードサマリー ---
        ws1 = wb.active
        ws1.title = "KWサマリー"

        summary_headers = [
            "No.", "キーワード", "スコア", "検索数",
            "通過数", "候補あり", "通過率(%)",
        ]
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col_idx, header in enumerate(summary_headers, 1):
            cell = ws1.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        sorted_data = sorted(session_data, key=lambda x: -x.get("score", 0))

        for i, kw_data in enumerate(sorted_data, 1):
            total_searched = kw_data.get("total_searched", 0)
            pass_count = kw_data.get("pass_count", 0)
            pass_rate = (pass_count / total_searched * 100) if total_searched > 0 else 0

            row = [
                i,
                kw_data["keyword"],
                kw_data.get("score", 0),
                total_searched,
                pass_count,
                len(kw_data.get("products", [])),
                round(pass_rate, 1),
            ]
            for col_idx, val in enumerate(row, 1):
                ws1.cell(row=i + 1, column=col_idx, value=val)

        # 列幅自動調整
        self._auto_column_width(ws1)

        # --- Sheet2: 全商品詳細 ---
        ws2 = wb.create_sheet(title="全商品詳細")

        detail_headers = [
            "No.", "重複", "候補状態", "キーワード", "KWスコア",
            "ASIN", "商品タイトル", "Amazon価格(円)", "BSR",
            "レビュー数", "評価", "FBA", "バリエ数",
            "月間販売数", "月間売上(円)", "セラー名",
            "1688商品名", "1688価格(元)", "最小ロット",
            "類似度(%)", "利益(円)", "利益率(%)",
            "総コスト(円)", "Amazon URL", "1688 URL",
        ]

        for col_idx, header in enumerate(detail_headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        row_idx = 2
        seen_asins: set[str] = set()
        dup_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        no_cand_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        for kw_data in sorted_data:
            kw = kw_data["keyword"]
            kw_score = kw_data.get("score", 0)

            # all_filtered があればそれを使い、なければ products にフォールバック
            display_products = kw_data.get("all_filtered") or kw_data.get("products", [])

            for prod_data in display_products:
                am = prod_data.get("amazon", {})
                asin = am.get("asin", "")
                # セッション内重複はスキップ（過去被りは記載する）
                is_dup = prod_data.get("is_duplicate", False)
                if asin and asin in seen_asins and not is_dup:
                    continue
                if asin:
                    seen_asins.add(asin)
                candidates = prod_data.get("candidates", [])
                has_candidates = len(candidates) > 0
                best = candidates[0] if candidates else {}
                profit_data = best.get("profit", {})
                alibaba_data = best.get("alibaba", {})

                # 候補状態
                if has_candidates:
                    candidate_status = "候補あり"
                else:
                    reason = prod_data.get("no_candidates_reason", "")
                    candidate_status = f"候補なし（{reason}）" if reason else "候補なし"

                row = [
                    row_idx - 1,
                    "重複" if is_dup else "",
                    candidate_status,
                    kw,
                    kw_score,
                    am.get("asin", ""),
                    (am.get("title", "") or "")[:80],
                    am.get("price", 0),
                    am.get("bsr", 0),
                    am.get("review_count", 0),
                    am.get("rating", ""),
                    "FBA" if am.get("is_fba") else "FBM",
                    am.get("variation_count", 1),
                    prod_data.get("estimated_monthly_sales", 0),
                    prod_data.get("estimated_monthly_revenue", 0),
                    am.get("seller_name", ""),
                    (alibaba_data.get("title", "") or "")[:60] if has_candidates else "",
                    alibaba_data.get("price_cny", 0) if has_candidates else "",
                    alibaba_data.get("min_order", "") if has_candidates else "",
                    round(best.get("combined_score", 0) * 100, 1) if best else "",
                    profit_data.get("profit", 0) if has_candidates else "",
                    profit_data.get("profit_rate_percentage", 0) if has_candidates else "",
                    profit_data.get("total_cost", 0) if has_candidates else "",
                    am.get("product_url", "") or f"https://www.amazon.co.jp/dp/{am.get('asin', '')}",
                    alibaba_data.get("product_url", "") if has_candidates else "",
                ]
                for col_idx, val in enumerate(row, 1):
                    cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                    if is_dup:
                        cell.fill = dup_fill
                        cell.font = Font(color="999999")
                    elif not has_candidates:
                        cell.fill = no_cand_fill
                        cell.font = Font(color="999999")
                row_idx += 1

        # 列幅自動調整
        self._auto_column_width(ws2)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = self.output_dir / f"session_{timestamp}.xlsx"
        wb.save(filepath)

        logger.info(f"セッション統合Excel生成: {filepath}")
        return str(filepath)

    @staticmethod
    def _auto_column_width(ws) -> None:
        """列幅を自動調整"""
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    val = str(cell.value)
                    length = sum(2 if ord(c) > 127 else 1 for c in val)
                    max_length = max(max_length, length)
            ws.column_dimensions[col_letter].width = min(max_length + 4, 60)


def _score_badge(score: float) -> str:
    """スコアに応じたバッジHTML"""
    if score >= 30:
        return '<span class="badge badge-3">&#9733;&#9733;&#9733;</span>'
    elif score >= 15:
        return '<span class="badge badge-2">&#9733;&#9733;</span>'
    elif score > 0:
        return '<span class="badge badge-1">&#9733;</span>'
    else:
        return '<span class="badge badge-0">&#9734;</span>'


_SESSION_HTML_TEMPLATE = r'''<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Auto Research Session Report</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI','Meiryo','Hiragino Sans',sans-serif;background:#f0f2f5;color:#333;padding:20px}
.dashboard{background:#fff;border-radius:12px;padding:24px;margin-bottom:24px;box-shadow:0 2px 8px rgba(0,0,0,.06)}
.dashboard h1{font-size:22px;color:#1a1a2e;margin-bottom:12px}
.dashboard .stats-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:16px;margin-bottom:16px}
.stat-card{background:#f8f9fa;border-radius:8px;padding:16px;text-align:center}
.stat-card .label{font-size:12px;color:#888;margin-bottom:4px}
.stat-card .value{font-size:24px;font-weight:bold;color:#1a1a2e}
.top-keywords{font-size:13px;color:#666;margin-top:12px}
.top-keywords span{display:inline-block;background:#e3f2fd;color:#1565c0;padding:2px 8px;border-radius:12px;margin:2px 4px;font-size:12px}

.kw-section{background:#fff;border-radius:12px;margin-bottom:16px;box-shadow:0 2px 8px rgba(0,0,0,.06);overflow:hidden}
.kw-header{padding:16px 20px;cursor:pointer;display:flex;align-items:center;gap:12px;border-bottom:1px solid #eee;user-select:none}
.kw-header:hover{background:#f8f9fa}
.kw-header .arrow{transition:transform .2s;font-size:12px;color:#888}
.kw-header .arrow.open{transform:rotate(90deg)}
.kw-header .kw-name{font-size:16px;font-weight:bold;color:#1a1a2e;flex:1}
.kw-header .kw-meta{font-size:12px;color:#888}
.badge{display:inline-block;padding:2px 8px;border-radius:10px;font-size:12px;font-weight:bold;margin-right:8px}
.badge-3{background:#c8e6c9;color:#2e7d32}
.badge-2{background:#fff9c4;color:#f57f17}
.badge-1{background:#ffccbc;color:#bf360c}
.badge-0{background:#f5f5f5;color:#999}
.kw-body{display:none;padding:16px 20px}
.kw-body.open{display:block}

.product-card{background:#fff;border-radius:12px;padding:24px;margin-bottom:24px;box-shadow:0 1px 4px rgba(0,0,0,.04)}
.product-num{display:inline-block;background:#1a1a2e;color:#fff;font-size:12px;font-weight:bold;padding:2px 10px;border-radius:12px;margin-bottom:12px}
.amazon-section{display:flex;gap:16px;padding-bottom:16px;border-bottom:1px solid #eee;margin-bottom:16px}
.amazon-img{width:130px;height:130px;object-fit:contain;border:1px solid #eee;border-radius:8px;background:#fafafa;flex-shrink:0}
.amazon-info{flex:1;min-width:0}
.amazon-info h3{font-size:14px;color:#333;margin-bottom:8px;line-height:1.5;display:-webkit-box;-webkit-line-clamp:2;-webkit-box-orient:vertical;overflow:hidden}
.amazon-price{font-size:20px;font-weight:bold;color:#B12704;margin-bottom:6px}
.amazon-details{font-size:12px;color:#666;line-height:1.8}
.amazon-details span{display:inline-block;margin-right:12px}
.amazon-link{font-size:12px;color:#0066c0;text-decoration:none;margin-top:4px;display:inline-block}
.amazon-link:hover{text-decoration:underline}

.candidates-label{font-size:13px;font-weight:bold;color:#555;margin-bottom:12px}
.candidates-row{display:flex;gap:12px;flex-wrap:wrap}
.candidate{border:2px solid #e0e0e0;border-radius:10px;padding:10px;width:175px;transition:all .15s;background:#fff}
.candidate:hover{border-color:#90CAF9;transform:translateY(-2px);box-shadow:0 4px 12px rgba(0,0,0,.1)}
.candidate-img{width:100%;height:140px;object-fit:contain;border-radius:6px;background:#fafafa;display:block}
.candidate-price{font-size:16px;font-weight:bold;color:#c62828;margin:8px 0 4px}
.candidate-profit{font-size:13px;font-weight:bold;margin-bottom:4px}
.candidate-profit.good{color:#2E7D32}
.candidate-profit.mid{color:#F57F17}
.candidate-profit.low{color:#c62828}
.candidate-detail{font-size:11px;color:#888;margin-bottom:4px}
.candidate-title{font-size:10px;color:#555;margin-bottom:4px;line-height:1.3;display:-webkit-box;-webkit-line-clamp:2;-webkit-box-orient:vertical;overflow:hidden}
.candidate-link{font-size:11px;color:#1565C0;text-decoration:none;word-break:break-all;display:block;margin-top:4px}
.candidate-link:hover{text-decoration:underline}
.candidate-shop{font-size:10px;color:#999;margin-top:2px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
</style>
</head>
<body>

<div class="dashboard" id="dashboard"></div>
<div id="kwSections"></div>

<script>
var DATA = __DATA_PLACEHOLDER__;

function init() {
  renderDashboard();
  renderKeywordSections();
}

function renderDashboard() {
  var s = DATA.stats;
  var seenAsins = {};
  var totalProducts = 0;
  var totalDuplicates = 0;
  DATA.keywords.forEach(function(kw) {
    kw.products.forEach(function(p) {
      var asin = p.amazon && p.amazon.asin;
      if (p.is_duplicate) { totalDuplicates++; return; }
      if (asin && !seenAsins[asin]) {
        seenAsins[asin] = true;
        totalProducts++;
      }
    });
  });

  var topKws = DATA.keywords.slice(0, 5);
  var topHtml = topKws.map(function(kw) {
    return '<span>' + esc(kw.keyword) + ' (' + kw.score + ')</span>';
  }).join('');

  document.getElementById('dashboard').innerHTML =
    '<h1>Auto Research Session Report</h1>' +
    '<div class="stats-grid">' +
      '<div class="stat-card"><div class="label">KW</div><div class="value">' + s.total_keywords + '</div></div>' +
      '<div class="stat-card"><div class="label">Candidate KW</div><div class="value">' + DATA.keywords.length + '</div></div>' +
      '<div class="stat-card"><div class="label">Products</div><div class="value">' + totalProducts + '</div></div>' +
      (totalDuplicates > 0 ? '<div class="stat-card"><div class="label">Duplicates</div><div class="value" style="color:#999">' + totalDuplicates + '</div></div>' : '') +
      '<div class="stat-card"><div class="label">Time</div><div class="value">' + esc(s.elapsed_str) + '</div></div>' +
    '</div>' +
    '<div class="top-keywords">Top: ' + topHtml + '</div>' +
    '<div style="font-size:12px;color:#999;margin-top:8px">' + esc(DATA.generated_at) + '</div>';
}

function renderKeywordSections() {
  var container = document.getElementById('kwSections');
  DATA.keywords.forEach(function(kwData, ki) {
    container.appendChild(buildKwSection(kwData, ki));
  });
}

function buildKwSection(kwData, ki) {
  var section = document.createElement('div');
  section.className = 'kw-section';

  var displayProducts = kwData.all_filtered || kwData.products;
  var withCandidates = displayProducts.filter(function(p) { return p.candidates && p.candidates.length > 0; }).length;
  var badge = scoreBadge(kwData.score);
  var headerHtml =
    '<div class="kw-header" onclick="toggleSection(this)">' +
      '<span class="arrow">&#9654;</span>' +
      badge +
      '<span class="kw-name">' + esc(kwData.keyword) + '</span>' +
      '<span class="kw-meta">Score: ' + kwData.score +
        ' | ' + kwData.pass_count + '/' + kwData.total_searched +
        ' | &#36890;&#36942;: ' + displayProducts.length +
        ' | &#20505;&#35036;&#12354;&#12426;: ' + withCandidates + '</span>' +
    '</div>';

  var bodyHtml = '<div class="kw-body">';
  displayProducts.forEach(function(prod, pi) {
    bodyHtml += buildProductCard(prod, pi);
  });
  bodyHtml += '</div>';

  section.innerHTML = headerHtml + bodyHtml;
  return section;
}

function buildProductCard(prod, pi) {
  var am = prod.amazon;
  var dims = am.dimensions ? am.dimensions[0]+'x'+am.dimensions[1]+'x'+am.dimensions[2]+'cm' : '-';
  var weight = am.weight_kg ? am.weight_kg+'kg' : '-';

  var isDup = prod.is_duplicate;
  var noCand = !prod.candidates || prod.candidates.length === 0;
  var cardStyle = isDup ? ' style="opacity:0.6;border-left:4px solid #e53935"' : noCand ? ' style="background:#fafafa;border-left:3px solid #ddd"' : '';
  var html = '<div class="product-card"' + cardStyle + '>';
  html += '<span class="product-num">#' + (pi+1) + '</span>';
  if (isDup) html += '<span style="background:#e53935;color:#fff;font-size:12px;font-weight:bold;padding:2px 10px;border-radius:12px;margin-left:8px">&#37325;&#35079;</span>';
  html += '<div class="amazon-section">';
  html += '<img class="amazon-img" src="' + esc(am.image_url) + '" onerror="this.style.display=\'none\'">';
  html += '<div class="amazon-info">';
  html += '<h3>' + esc(am.title) + '</h3>';
  html += '<div class="amazon-price">&yen;' + num(am.price) + '</div>';
  html += '<div class="amazon-details">';
  html += '<span>ASIN: ' + am.asin + '</span>';
  html += '<span>BSR: ' + num(am.bsr) + '</span>';
  html += '<span>&#12524;&#12499;&#12517;&#12540;: ' + am.review_count + '&#20214;</span>';
  if (am.rating) html += '<span>&#9733;' + am.rating + '</span>';
  html += '<span>' + (am.is_fba ? 'FBA' : 'FBM') + '</span>';
  html += '<span>&#12496;&#12522;&#12456;: ' + (am.variation_count || 1) + '</span>';
  html += '<span>&#23544;&#27861;: ' + dims + '</span>';
  html += '<span>&#37325;&#37327;: ' + weight + '</span>';
  html += '</div>';
  html += '<div class="amazon-details"><span>&#26376;&#38291;&#36009;&#22770;: ' + num(prod.estimated_monthly_sales) + '&#20491;</span>';
  html += '<span>&#26376;&#38291;&#22770;&#19978;: &yen;' + num(prod.estimated_monthly_revenue) + '</span></div>';
  if (am.seller_name) html += '<div class="amazon-details"><span>&#12475;&#12521;&#12540;: ' + esc(am.seller_name) + '</span></div>';
  html += '<a class="amazon-link" href="' + esc(am.product_url || ('https://www.amazon.co.jp/dp/' + am.asin)) + '" target="_blank">Amazon &#12391;&#35211;&#12427; &rarr;</a>';
  html += '</div></div>';

  // Candidates
  var cands = prod.candidates || [];
  if (cands.length > 0) {
    html += '<div class="candidates-label">1688&#20505;&#35036; (' + cands.length + '&#20214;)</div>';
    html += '<div class="candidates-row">';
    cands.forEach(function(cand) {
      var profitPct = cand.profit.profit_rate_percentage || 0;
      var profitClass = profitPct >= 25 ? 'good' : profitPct >= 15 ? 'mid' : 'low';
      html += '<div class="candidate">';
      html += '<img class="candidate-img" src="' + esc(cand.alibaba.image_url) + '" onerror="this.style.display=\'none\'">';
      if (cand.alibaba.title) html += '<div class="candidate-title">' + esc(cand.alibaba.title) + '</div>';
      html += '<div class="candidate-price">' + cand.alibaba.price_cny + '&#20803;</div>';
      html += '<div class="candidate-profit ' + profitClass + '">&#21033;&#30410;: &yen;' + num(cand.profit.profit) + ' (' + profitPct + '%)</div>';
      html += '<div class="candidate-detail">&#32207;&#12467;&#12473;&#12488;: &yen;' + num(cand.profit.total_cost) + '</div>';
      html += '<div class="candidate-detail">&#39006;&#20284;&#24230;: ' + ((cand.combined_score || 0) * 100).toFixed(1) + '%</div>';
      if (cand.alibaba.min_order) html += '<div class="candidate-detail">&#26368;&#23567;&#12525;&#12483;&#12488;: ' + cand.alibaba.min_order + '</div>';
      if (cand.alibaba.product_url) html += '<a class="candidate-link" href="' + esc(cand.alibaba.product_url) + '" target="_blank">1688 &rarr;</a>';
      if (cand.alibaba.shop_name) html += '<div class="candidate-shop">' + esc(cand.alibaba.shop_name) + '</div>';
      html += '</div>';
    });
    html += '</div>';
  } else {
    html += '<div style="color:#999;font-size:13px;padding:16px;text-align:center;background:#f9f9f9;border-radius:8px;border:1px dashed #ddd">1688&#20505;&#35036;&#12394;&#12375;';
    if (prod.no_candidates_reason) {
      html += ' &mdash; ' + esc(prod.no_candidates_reason);
    }
    html += '</div>';
  }

  html += '</div>';
  return html;
}

function toggleSection(header) {
  var arrow = header.querySelector('.arrow');
  var body = header.nextElementSibling;
  if (body.classList.contains('open')) {
    body.classList.remove('open');
    arrow.classList.remove('open');
  } else {
    body.classList.add('open');
    arrow.classList.add('open');
  }
}

function scoreBadge(score) {
  if (score >= 30) return '<span class="badge badge-3">&#9733;&#9733;&#9733;</span>';
  if (score >= 15) return '<span class="badge badge-2">&#9733;&#9733;</span>';
  if (score > 0)  return '<span class="badge badge-1">&#9733;</span>';
  return '<span class="badge badge-0">&#9734;</span>';
}

function esc(s) {
  if (!s) return '';
  return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}

function num(n) {
  if (n == null) return '-';
  return Number(n).toLocaleString();
}

init();
</script>
</body>
</html>'''
